import pandas as pd
import csv
import os
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference
from collections import defaultdict, Counter
import re

# File paths
input_tsv = r"C:\Users\ebarthel3\Desktop\STING 7.0\test-file.tsv"
output_file = r"C:\Users\ebarthel3\Desktop\STING 7.0\fOutputAndaReport.xlsx"

def sanitize_text(text):
    """Remove or replace Unicode characters that aren't supported by core fonts"""
    if not isinstance(text, str):
        return str(text)
    
    # Replace common problematic characters
    replacements = {
        '—': '-',      # em-dash
        '–': '-',      # en-dash
        '"': '"',     # left double quote
        '"': '"',     # right double quote
        ''': "'",    # right single quote
        ''': "'",    # left single quote
        '…': '...',    # ellipsis
    }
    
    for char, replacement in replacements.items():
        text = text.replace(char, replacement)
    
    # Remove any remaining non-latin-1 characters
    text = text.encode('latin-1', errors='ignore').decode('latin-1')
    
    return text

def extract_themes(text_list, max_themes=5):
    """Extract common themes/words from a list of text responses"""
    if not text_list:
        return []
    
    # Stop words to exclude
    stop_words = {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 
                  'of', 'is', 'are', 'be', 'been', 'being', 'have', 'has', 'had', 'do', 
                  'does', 'did', 'will', 'would', 'could', 'should', 'may', 'might', 'must',
                  'i', 'you', 'he', 'she', 'it', 'we', 'they', 'me', 'him', 'her', 'us', 'them',
                  'this', 'that', 'these', 'those', 'my', 'your', 'his', 'her', 'its', 'our', 'their',
                  'from', 'by', 'with', 'as', 'if', 'no', 'not', 'more', 'most', 'other', 'than',
                  'very', 'so', 'too', 'any', 'all', 'each', 'every', 'both', 'few', 'some',
                  'such', 'what', 'which', 'who', 'whom', 'why', 'where', 'when', 'how', 'or'}
    
    # Extract words
    all_words = []
    for text in text_list:
        if text and text.strip() and text.strip().upper() != 'N/A' and '[NO RESPONSE]' not in text.upper():
            words = re.findall(r'\b[a-z]+\b', text.lower())
            all_words.extend([w for w in words if w not in stop_words and len(w) > 3])
    
    # Count word frequency
    word_counts = Counter(all_words)
    return word_counts.most_common(max_themes)

def extract_expertise_areas(background_list):
    """Extract expertise categories from background text"""
    expertise_keywords = {
        'AI/Machine Learning': ['ai', 'ml', 'machine learning', 'artificial intelligence', 'deep learning', 'neural', 'llm'],
        'Engineering': ['engineering', 'engineer', 'aerospace', 'systems', 'electrical', 'mechanical', 'software', 'hardware'],
        'Design (HCD/UX)': ['design', 'human-centered', 'ux', 'user experience', 'industrial', 'hcd'],
        'Cybersecurity': ['cybersecurity', 'security', 'cyber', 'encryption'],
        'Data Science': ['data', 'analytics', 'analysis', 'database', 'statistical'],
        'Research': ['research', 'researcher'],
        'Leadership/Management': ['manager', 'lead', 'officer', 'director', 'management', 'leadership'],
        'Military': ['military', 'marine', 'army', 'navy', 'air force', 'infantry', 'commissioned'],
        'Policy/Government': ['policy', 'government', 'federal', 'political'],
    }
    
    expertise_counts = defaultdict(int)
    
    for text in background_list:
        if text and text.strip() and '[NO RESPONSE]' not in text.upper():
            text_lower = text.lower()
            for category, keywords in expertise_keywords.items():
                for keyword in keywords:
                    if keyword in text_lower:
                        expertise_counts[category] += 1
                        break  # Count each category only once per applicant
    
    return expertise_counts

def parse_qualtrics_tsv(file_path):
    """
    Parse Qualtrics TSV export format:
    - Row 0: Question IDs (Q3, Q20, etc.)
    - Row 1: Full question text
    - Row 2: Import metadata (skip)
    - Row 3+: Applicant responses (one row per applicant)
    """
    
    # Rating scale mappings
    familiarity_scale = {
        '1': 'Not familiar at all',
        '2': 'Slightly familiar',
        '3': 'Moderately familiar',
        '4': 'Very familiar',
        '5': 'Extremely familiar'
    }
    
    with open(file_path, 'r', encoding='utf-16') as f:
        reader = csv.reader(f, delimiter='\t')
        lines = list(reader)
    
    if len(lines) < 4:
        print("Error: TSV file doesn't have enough rows")
        return None, None, None
    
    question_ids = lines[0]
    question_texts = lines[1]
    # Skip row 2 (import metadata)
    data_rows = lines[3:]  # Start from row 3
    
    # Find the starting column for questions (after metadata columns)
    # Typically starts at column 17 with Q3
    question_start_col = None
    for i, qid in enumerate(question_ids):
        if qid.startswith('Q') and not qid.startswith('Q_'):
            question_start_col = i
            break
    
    if question_start_col is None:
        print("Error: Could not find question columns")
        return None, None, None
    
    # Extract questions (exclude Source and supervisor email)
    questions = []
    for i in range(question_start_col, len(question_texts)):
        qid = question_ids[i]
        qtext = question_texts[i]
        
        # Skip metadata fields
        if qid in ['Source', 'Q40']:  # Q40 is supervisor email
            continue
        
        # Skip empty columns
        if not qtext or not qtext.strip():
            continue
        
        questions.append({
            'id': qid,
            'text': qtext,
            'column_index': i
        })
    
    print(f"Found {len(data_rows)} applicants")
    print(f"Found {len(questions)} questions")
    
    # Build applicant dictionary
    applicants = {}
    labs = defaultdict(list)
    
    for row in data_rows:
        if len(row) <= question_start_col:
            continue
        
        # Get applicant name (usually first question)
        name_col = questions[0]['column_index']
        applicant_name = row[name_col] if len(row) > name_col else ""
        
        if not applicant_name or not applicant_name.strip():
            continue
        
        applicant_name = applicant_name.strip()
        applicants[applicant_name] = {}
        
        # Extract lab/unit from column 19 (Q4 - "Which unit are you part of?")
        lab_col = 19
        lab = row[lab_col] if len(row) > lab_col else "Unknown"
        if lab and lab.strip():
            labs[lab.strip()].append(applicant_name)
        else:
            labs["Unknown"].append(applicant_name)
        
        # Extract all responses for this applicant
        for q in questions:
            col_idx = q['column_index']
            response = row[col_idx] if len(row) > col_idx else ""
            
            # Handle empty responses
            if not response or not response.strip():
                response = "[No response]"
            else:
                response = response.strip()
                
                # Map numeric ratings to descriptive text for familiarity questions (Q25_1, Q25_2)
                if q['id'] in ['Q25_1', 'Q25_2'] and response in familiarity_scale:
                    response = familiarity_scale[response]
                
                response = sanitize_text(response)
            
            applicants[applicant_name][q['text']] = response
    
    return applicants, questions, dict(labs)

def create_summary_sheet(writer, applicants, questions, labs):
    """Create comprehensive summary sheet with all analyses"""
    summary_data = {
        'Metric': [
            'Total Applicants',
            'Total Questions',
            'Report Generated'
        ],
        'Value': [
            len(applicants),
            len(questions),
            pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=0)
    
    worksheet = writer.sheets['Summary']
    worksheet.column_dimensions['A'].width = 40
    worksheet.column_dimensions['B'].width = 60
    
    for cell in worksheet[1]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    startrow = 6
    
    # ===== UNIT/LAB BREAKDOWN =====
    ws_title = worksheet[f'A{startrow}']
    ws_title.value = 'Unit/Lab Participation Breakdown'
    ws_title.font = Font(bold=True, size=12, color="FFFFFF")
    ws_title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    startrow += 1
    
    headers = ['Lab/Unit', 'Count', 'Applicants']
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=startrow, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    startrow += 1
    lab_start = startrow
    
    for lab in sorted(labs.keys(), key=lambda x: len(labs[x]), reverse=True):
        applicant_list = labs[lab]
        worksheet.cell(row=startrow, column=1).value = lab
        worksheet.cell(row=startrow, column=2).value = len(applicant_list)
        worksheet.cell(row=startrow, column=3).value = ', '.join(sorted(applicant_list))
        worksheet.cell(row=startrow, column=3).alignment = Alignment(wrap_text=True)
        startrow += 1
    
    lab_end = startrow - 1
    
    # Pie chart for units
    pie = PieChart()
    pie.title = "Unit Distribution"
    pie.style = 10
    data = Reference(worksheet, min_col=2, min_row=lab_start - 1, max_row=lab_end)
    labels = Reference(worksheet, min_col=1, min_row=lab_start, max_row=lab_end)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    worksheet.add_chart(pie, f"E{lab_start}")
    
    startrow += 2
    
    # ===== EXPERIENCE LEVEL DISTRIBUTION =====
    exp_title = worksheet[f'A{startrow}']
    exp_title.value = 'Experience Level Distribution'
    exp_title.font = Font(bold=True, size=12, color="FFFFFF")
    exp_title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    startrow += 1
    
    exp_question = None
    for q in questions:
        if q['id'] == 'Q24':
            exp_question = q['text']
            break
    
    if exp_question:
        exp_counts = defaultdict(int)
        exp_order = ['Entry level (0-2 years)', 'Novice (2-5 years)', 'Intermediate (5-10 years)', 'Advanced (10-15 years)', 'Expert (15+ years)']
        
        for applicant_name in applicants.keys():
            exp_level = applicants[applicant_name].get(exp_question, '')
            if exp_level and exp_level != '[No response]':
                exp_counts[exp_level] += 1
        
        worksheet.cell(row=startrow, column=1).value = 'Experience Level'
        worksheet.cell(row=startrow, column=2).value = 'Count'
        for cell in [worksheet.cell(row=startrow, column=1), worksheet.cell(row=startrow, column=2)]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        startrow += 1
        exp_data_start = startrow
        
        for exp_level in exp_order:
            worksheet.cell(row=startrow, column=1).value = exp_level
            worksheet.cell(row=startrow, column=2).value = exp_counts.get(exp_level, 0)
            startrow += 1
        
        exp_data_end = startrow - 1
        
        # Bar chart
        exp_chart = BarChart()
        exp_chart.type = "col"
        exp_chart.style = 10
        exp_chart.title = "Experience Distribution"
        data = Reference(worksheet, min_col=2, min_row=exp_data_start - 1, max_row=exp_data_end)
        cats = Reference(worksheet, min_col=1, min_row=exp_data_start, max_row=exp_data_end)
        exp_chart.add_data(data, titles_from_data=True)
        exp_chart.set_categories(cats)
        worksheet.add_chart(exp_chart, f"E{exp_data_start}")
    
    startrow += 2
    
    # ===== FAMILIARITY RATINGS =====
    fam_title = worksheet[f'A{startrow}']
    fam_title.value = 'Familiarity Ratings Analysis'
    fam_title.font = Font(bold=True, size=12, color="FFFFFF")
    fam_title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    startrow += 1
    
    fam_scale_reverse = {
        'Not familiar at all': 1,
        'Slightly familiar': 2,
        'Moderately familiar': 3,
        'Very familiar': 4,
        'Extremely familiar': 5
    }
    
    for q in questions:
        if q['id'] in ['Q25_1', 'Q25_2']:
            scores = []
            for applicant_name in applicants.keys():
                response = applicants[applicant_name].get(q['text'], '')
                if response and response != '[No response]':
                    score = fam_scale_reverse.get(response, None)
                    if score:
                        scores.append(score)
            
            if scores:
                avg_score = sum(scores) / len(scores)
                label = q['text'][:60]
                worksheet.cell(row=startrow, column=1).value = f"{label} (Avg: {avg_score:.2f}/5)"
                startrow += 1
    
    startrow += 1
    
    # ===== EXPERIENCE BY LAB MATRIX =====
    matrix_title = worksheet[f'A{startrow}']
    matrix_title.value = 'Experience by Lab (Cross-Tab)'
    matrix_title.font = Font(bold=True, size=12, color="FFFFFF")
    matrix_title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    startrow += 1
    
    if exp_question:
        exp_levels_short = ['Entry', 'Novice', 'Intermediate', 'Advanced', 'Expert']
        worksheet.cell(row=startrow, column=1).value = 'Lab'
        for i, exp in enumerate(exp_levels_short, 2):
            worksheet.cell(row=startrow, column=i).value = exp
        
        for cell in worksheet.iter_rows(min_row=startrow, max_row=startrow, min_col=1, max_col=6):
            for c in cell:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        startrow += 1
        
        for lab in sorted(labs.keys()):
            worksheet.cell(row=startrow, column=1).value = lab
            for i, exp_full in enumerate(exp_order, 2):
                count = sum(1 for applicant in labs[lab] if applicants[applicant].get(exp_question, '') == exp_full)
                worksheet.cell(row=startrow, column=i).value = count
            startrow += 1
    
    startrow += 2
    
    # ===== WORKSHOP ATTENDANCE =====
    workshop_title = worksheet[f'A{startrow}']
    workshop_title.value = 'Workshop Attendance'
    workshop_title.font = Font(bold=True, size=12, color="FFFFFF")
    workshop_title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    startrow += 1
    
    workshop_question = None
    for q in questions:
        if q['id'] == 'Q30':
            workshop_question = q['text']
            break
    
    if workshop_question:
        full_attendance = 0
        conflicts = 0
        
        for applicant_name in applicants.keys():
            response = applicants[applicant_name].get(workshop_question, '')
            if not response or response.strip().upper() in ['N/A', '[NO RESPONSE]', '']:
                full_attendance += 1
            else:
                conflicts += 1
        
        worksheet.cell(row=startrow, column=1).value = 'Attendance Type'
        worksheet.cell(row=startrow, column=2).value = 'Count'
        for cell in [worksheet.cell(row=startrow, column=1), worksheet.cell(row=startrow, column=2)]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        startrow += 1
        att_start = startrow
        
        worksheet.cell(row=startrow, column=1).value = 'Can attend all'
        worksheet.cell(row=startrow, column=2).value = full_attendance
        startrow += 1
        
        worksheet.cell(row=startrow, column=1).value = 'Has conflicts'
        worksheet.cell(row=startrow, column=2).value = conflicts
        att_end = startrow
        
        # Pie chart
        att_pie = PieChart()
        att_pie.title = "Workshop Attendance"
        att_data = Reference(worksheet, min_col=2, min_row=att_start - 1, max_row=att_end)
        att_labels = Reference(worksheet, min_col=1, min_row=att_start, max_row=att_end)
        att_pie.add_data(att_data, titles_from_data=True)
        att_pie.set_categories(att_labels)
        worksheet.add_chart(att_pie, f"E{att_start}")
    
    startrow += 3
    
    # ===== QUALITATIVE ANALYSIS =====
    qual_title = worksheet[f'A{startrow}']
    qual_title.value = 'Qualitative Analysis: Themes & Patterns'
    qual_title.font = Font(bold=True, size=12, color="FFFFFF")
    qual_title.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    startrow += 1
    
    # Common Challenges (Q33)
    worksheet.cell(row=startrow, column=1).value = 'Common Challenges (from Q33)'
    worksheet.cell(row=startrow, column=1).font = Font(bold=True, size=11)
    startrow += 1
    
    challenges_question = None
    for q in questions:
        if q['id'] == 'Q33':
            challenges_question = q['text']
            break
    
    if challenges_question:
        challenges_responses = [applicants[name].get(challenges_question, '') for name in applicants.keys() 
                               if applicants[name].get(challenges_question, '') and '[No response]' not in applicants[name].get(challenges_question, '')]
        challenges = extract_themes(challenges_responses, max_themes=5)
        if challenges:
            for theme, count in challenges:
                worksheet.cell(row=startrow, column=1).value = f"  • {theme.capitalize()} ({count}x)"
                startrow += 1
        else:
            worksheet.cell(row=startrow, column=1).value = "  • No themes extracted"
            startrow += 1
    
    startrow += 1
    
    # Motivations/Goals (Q21)
    worksheet.cell(row=startrow, column=1).value = 'Top Motivations/Goals (from Q21)'
    worksheet.cell(row=startrow, column=1).font = Font(bold=True, size=11)
    startrow += 1
    
    motivations_question = None
    for q in questions:
        if q['id'] == 'Q21':
            motivations_question = q['text']
            break
    
    if motivations_question:
        motivations_responses = [applicants[name].get(motivations_question, '') for name in applicants.keys()
                                if applicants[name].get(motivations_question, '') and '[No response]' not in applicants[name].get(motivations_question, '')]
        motivations = extract_themes(motivations_responses, max_themes=5)
        if motivations:
            for theme, count in motivations:
                worksheet.cell(row=startrow, column=1).value = f"  • {theme.capitalize()} ({count}x)"
                startrow += 1
        else:
            worksheet.cell(row=startrow, column=1).value = "  • No themes extracted"
            startrow += 1
    
    startrow += 1
    
    # Why Select You - Key Strengths (Q18)
    worksheet.cell(row=startrow, column=1).value = 'Key Strengths Cited (from Q18)'
    worksheet.cell(row=startrow, column=1).font = Font(bold=True, size=11)
    startrow += 1
    
    selection_question = None
    for q in questions:
        if q['id'] == 'Q18':
            selection_question = q['text']
            break
    
    if selection_question:
        selection_responses = [applicants[name].get(selection_question, '') for name in applicants.keys()
                              if applicants[name].get(selection_question, '') and '[No response]' not in applicants[name].get(selection_question, '')]
        strengths = extract_themes(selection_responses, max_themes=6)
        if strengths:
            for theme, count in strengths:
                worksheet.cell(row=startrow, column=1).value = f"  • {theme.capitalize()} ({count}x)"
                startrow += 1
        else:
            worksheet.cell(row=startrow, column=1).value = "  • No themes extracted"
            startrow += 1
    
    startrow += 2
    
    # ===== CAPABILITY INVENTORY =====
    inv_title = worksheet[f'A{startrow}']
    inv_title.value = 'Capability Inventory'
    inv_title.font = Font(bold=True, size=12, color="FFFFFF")
    inv_title.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    startrow += 1
    
    # Expertise Areas (Q22)
    worksheet.cell(row=startrow, column=1).value = 'Expertise Areas (from Q22)'
    worksheet.cell(row=startrow, column=1).font = Font(bold=True, size=11)
    startrow += 1
    
    background_question = None
    for q in questions:
        if q['id'] == 'Q22':
            background_question = q['text']
            break
    
    if background_question:
        background_responses = [applicants[name].get(background_question, '') for name in applicants.keys()
                               if applicants[name].get(background_question, '') and '[No response]' not in applicants[name].get(background_question, '')]
        expertise = extract_expertise_areas(background_responses)
        if expertise:
            for category in sorted(expertise.keys(), key=lambda x: expertise[x], reverse=True):
                count = expertise[category]
                worksheet.cell(row=startrow, column=1).value = f"  • {category}: {count} applicant(s)"
                startrow += 1
        else:
            worksheet.cell(row=startrow, column=1).value = "  • No expertise areas identified"
            startrow += 1
    
    startrow += 1
    
    # Military/Leadership Experience
    worksheet.cell(row=startrow, column=1).value = 'Military/Leadership Experience'
    worksheet.cell(row=startrow, column=1).font = Font(bold=True, size=11)
    startrow += 1
    
    military_keywords = ['military', 'marine', 'army', 'navy', 'officer', 'infantry', 'manager', 'lead', 'leadership', 'director', 'commissioned']
    military_applicants = []
    
    if background_question and selection_question:
        for name in applicants.keys():
            text = (applicants[name].get(background_question, '') + " " + 
                   applicants[name].get(selection_question, '')).lower()
            if any(keyword in text for keyword in military_keywords):
                military_applicants.append(name)
    
    if military_applicants:
        for name in sorted(military_applicants):
            worksheet.cell(row=startrow, column=1).value = f"  • {name}"
            startrow += 1
    else:
        worksheet.cell(row=startrow, column=1).value = "  • None identified"
        startrow += 1
    
    startrow += 1
    
    # Research Focus Areas
    worksheet.cell(row=startrow, column=1).value = 'Research Focus Areas'
    worksheet.cell(row=startrow, column=1).font = Font(bold=True, size=11)
    startrow += 1
    
    research_keywords = {
        'AI/ML': ['ai', 'machine learning', 'neural', 'deep learning', 'llm'],
        'Human-Centered Design': ['human-centered', 'hcd', 'user research', 'user experience'],
        'Systems Engineering': ['systems', 'engineering', 'integration'],
        'Policy': ['policy', 'governance', 'political'],
        'Sustainability': ['sustainability', 'environment', 'climate'],
        'Social Impact': ['social', 'community', 'equity', 'public']
    }
    
    research_counts = defaultdict(int)
    
    for name in applicants.keys():
        bg_text = (applicants[name].get(background_question, '') + " " + 
                  applicants[name].get(selection_question, '')).lower()
        for category, keywords in research_keywords.items():
            for keyword in keywords:
                if keyword in bg_text:
                    research_counts[category] += 1
                    break
    
    if research_counts:
        for category in sorted(research_counts.keys(), key=lambda x: research_counts[x], reverse=True):
            if research_counts[category] > 0:
                worksheet.cell(row=startrow, column=1).value = f"  • {category}: {research_counts[category]} applicant(s)"
                startrow += 1
    else:
        worksheet.cell(row=startrow, column=1).value = "  • No research focus identified"
        startrow += 1

def create_applicant_sheets(writer, applicants):
    """Create individual sheets for each applicant with all their responses"""
    
    for applicant_name in sorted(applicants.keys()):
        responses = applicants[applicant_name]
        
        # Create DataFrame from applicant responses
        data = {
            'Question': list(responses.keys()),
            'Response': list(responses.values())
        }
        applicant_df = pd.DataFrame(data)
        
        # Sanitize sheet name - Excel doesn't allow: [ ] : * ? / \
        # Also max 31 characters
        sheet_name = applicant_name[:31]
        invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '')
        
        # Write to sheet
        applicant_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
        
        # Format the sheet
        worksheet = writer.sheets[sheet_name]
        
        # Format header row
        for cell in worksheet[1]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="top")
        
        # Auto-adjust column widths and wrap text
        for col in worksheet.columns:
            max_length = 0
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            worksheet.column_dimensions[col[0].column_letter].width = adjusted_width
        
        # Wrap text for responses and set row heights
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            # Set minimum row height
            worksheet.row_dimensions[row[0].row].height = 30

def process_tsv(input_tsv, output_file):
    print("Parsing Qualtrics TSV export...")
    applicants, questions, labs = parse_qualtrics_tsv(input_tsv)
    
    if not applicants:
        print("No applicant data found")
        return
    
    print(f"\nProcessing {len(applicants)} applicants with {len(questions)} questions each")
    print("\nQuestions included:")
    for i, q in enumerate(questions, 1):
        print(f"  {i}. {q['text'][:80]}{'...' if len(q['text']) > 80 else ''}")
    
    # Delete old file if it exists to avoid permission errors
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
            print(f"\nRemoved old file: {output_file}")
        except PermissionError:
            print(f"\nWarning: Could not delete old file (it may be open). Trying to create backup...")
            backup_file = output_file.replace('.xlsx', '_backup.xlsx')
            try:
                os.rename(output_file, backup_file)
                print(f"Created backup: {backup_file}")
            except:
                print("Error: Cannot write to file. Please close it if it's open in Excel.")
                return
    
    # Create Excel report
    print("\nCreating Excel report with comprehensive analyses...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        create_summary_sheet(writer, applicants, questions, labs)
        create_applicant_sheets(writer, applicants)
    
    print(f"\nExcel report generated: {output_file}")
    print(f"  - Summary sheet: 1 (with all analyses)")
    print(f"  - Applicant sheets: {len(applicants)}")
    print(f"  - Questions per applicant: {len(questions)}")

# Run the script
if __name__ == "__main__":
    process_tsv(input_tsv, output_file)
