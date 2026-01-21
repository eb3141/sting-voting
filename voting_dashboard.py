import streamlit as st
import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import hashlib

# Page config
st.set_page_config(page_title="STING Applicant Voting", layout="wide")

# ===== PASSWORD AUTHENTICATION =====
def check_password():
    """Returns `True` if the user had the correct password."""
    
    def password_entered():
        """Checks whether a password entered by the user is correct."""
        correct_password = st.secrets.get("voting_password", "sting2026")
        if st.session_state["password"] == correct_password:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # don't store password
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # First run, show input for password.
        st.markdown("""
        <div style='text-align: center; padding: 50px;'>
            <h1>🗳️ STING Applicant Voting Dashboard</h1>
            <p style='font-size: 18px; color: #666;'>Enter the voting password to access the system</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.text_input(
            "Voting Password:",
            type="password",
            on_change=password_entered,
            key="password",
            placeholder="Enter password"
        )
        st.info("📌 Contact your administrator if you don't have the password.")
        st.stop()  # Do not continue if check_password is not True.

    elif not st.session_state["password_correct"]:
        # Password not correct, show error.
        st.error("❌ Incorrect password. Please try again.")
        st.text_input(
            "Voting Password:",
            type="password",
            on_change=password_entered,
            key="password",
            placeholder="Enter password"
        )
        st.stop()  # Do not continue if check_password is not True.

# Check password before showing app
check_password()

# ===== MAIN APP =====
st.title("🗳️ STING Applicant Voting Dashboard")

# File paths (use relative paths for Streamlit Cloud compatibility)
votes_file = "votes.csv"
excel_file = "fOutputAndaReport.xlsx"
export_file = "Voting_Results_Summary.xlsx"

# Check if Excel file exists
if not os.path.exists(excel_file):
    st.error(f"❌ Error: {excel_file} not found!")
    st.info("📌 Please ensure 'fOutputAndaReport.xlsx' is uploaded to the GitHub repository.")
    st.stop()

# Initialize votes CSV if it doesn't exist
if not os.path.exists(votes_file):
    df_votes = pd.DataFrame(columns=[
        'timestamp', 'judge_name', 'applicant_name', 'status', 
        'rating', 'comment', 'original_status', 'original_rating', 'vote_version'
    ])
    df_votes.to_csv(votes_file, index=False)

# Load applicants from Excel
@st.cache_data
def load_applicants():
    wb = openpyxl.load_workbook(excel_file)
    applicants = {}
    
    # Get all applicant sheet names (skip Summary sheet)
    for sheet_name in wb.sheetnames:
        if sheet_name != 'Summary':
            ws = wb[sheet_name]
            applicant_data = {}
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                question = row[0].value
                response = row[1].value
                if question:
                    applicant_data[question] = response if response else "[No response]"
            applicants[sheet_name] = applicant_data
    
    return applicants

# Load votes from CSV
def load_votes():
    if os.path.exists(votes_file):
        return pd.read_csv(votes_file)
    return pd.DataFrame(columns=[
        'timestamp', 'judge_name', 'applicant_name', 'status', 
        'rating', 'comment', 'original_status', 'original_rating', 'vote_version'
    ])

# Save vote to CSV
def save_vote(judge_name, applicant_name, status, rating, comment, original_status, original_rating):
    df_votes = load_votes()
    
    # Check if this judge already voted for this applicant
    existing = df_votes[(df_votes['judge_name'] == judge_name) & 
                        (df_votes['applicant_name'] == applicant_name)]
    
    if not existing.empty:
        # This is a revision - mark original
        original_status = existing.iloc[-1]['status']
        original_rating = existing.iloc[-1]['rating']
        vote_version = existing.iloc[-1]['vote_version'] + 1
    else:
        vote_version = 1
    
    new_vote = pd.DataFrame([{
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'judge_name': judge_name,
        'applicant_name': applicant_name,
        'status': status,
        'rating': rating,
        'comment': comment,
        'original_status': original_status,
        'original_rating': original_rating,
        'vote_version': vote_version
    }])
    
    df_votes = pd.concat([df_votes, new_vote], ignore_index=True)
    df_votes.to_csv(votes_file, index=False)
    st.success(f"✅ Vote saved for {applicant_name}")

# Get latest vote for applicant from a judge
def get_judge_vote(judge_name, applicant_name):
    df_votes = load_votes()
    judge_votes = df_votes[(df_votes['judge_name'] == judge_name) & 
                           (df_votes['applicant_name'] == applicant_name)]
    
    if judge_votes.empty:
        return None
    return judge_votes.iloc[-1]

# Get all votes for an applicant
def get_applicant_votes(applicant_name):
    df_votes = load_votes()
    return df_votes[df_votes['applicant_name'] == applicant_name]

# Create tabs
tab1, tab2, tab3 = st.tabs(["🗳️ Vote", "📊 Results Dashboard", "📥 Export Results"])

applicants = load_applicants()
applicant_names = sorted(list(applicants.keys()))

# ===== TAB 1: VOTING INTERFACE =====
with tab1:
    st.header("Submit Your Votes")
    
    # Judge name input
    col1, col2 = st.columns([2, 3])
    with col1:
        judge_name = st.text_input("👤 Enter your name:", key="judge_name_input", placeholder="e.g., John Smith")
    
    if not judge_name:
        st.warning("⚠️ Please enter your name to continue voting")
        st.stop()
    
    st.divider()
    
    # Display all applicants for voting
    for applicant_name in applicant_names:
        with st.expander(f"📋 {applicant_name}", expanded=False):
            # Get applicant details
            details = applicants[applicant_name]
            
            # Display applicant info
            info_cols = st.columns(3)
            with info_cols[0]:
                st.write("**Unit/Lab:**")
                for q, r in details.items():
                    if "unit" in q.lower():
                        st.write(r)
                        break
            
            with info_cols[1]:
                st.write("**Experience:**")
                for q, r in details.items():
                    if "experience level" in q.lower():
                        st.write(r)
                        break
            
            with info_cols[2]:
                st.write("**Background:**")
                for q, r in details.items():
                    if "background" in q.lower():
                        st.write(r[:100] + "..." if len(str(r)) > 100 else r)
                        break
            
            st.markdown("---")
            
            # Show other judges' votes for this applicant
            applicant_votes = get_applicant_votes(applicant_name)
            if not applicant_votes.empty:
                st.subheader("📝 Other Judges' Votes & Comments:")
                for idx, vote in applicant_votes.iterrows():
                    if vote['judge_name'] != judge_name:
                        status_emoji = {"Approve": "✅", "Reject": "❌", "Maybe": "❓"}
                        emoji = status_emoji.get(vote['status'], "")
                        st.write(f"**{vote['judge_name']}** {emoji} {vote['status']} | ⭐ {int(vote['rating'])}/5")
                        if pd.notna(vote['comment']) and vote['comment']:
                            st.write(f"_Comment: {vote['comment']}_")
                        if pd.notna(vote['original_status']) and vote['original_status']:
                            st.caption(f"Original vote: {vote['original_status']} | ⭐ {int(vote['original_rating'])}/5")
                st.divider()
            
            # Get current judge's vote if exists
            current_vote = get_judge_vote(judge_name, applicant_name)
            
            st.subheader("🗳️ Your Vote:")
            
            # Voting columns
            vote_col1, vote_col2 = st.columns(2)
            
            with vote_col1:
                status = st.radio(
                    "Status:",
                    options=["Approve", "Reject", "Maybe"],
                    key=f"status_{applicant_name}",
                    index=["Approve", "Reject", "Maybe"].index(current_vote['status']) if current_vote is not None else 0
                )
            
            with vote_col2:
                rating = st.slider(
                    "Rating (1-5):",
                    min_value=1,
                    max_value=5,
                    value=int(current_vote['rating']) if current_vote is not None else 3,
                    key=f"rating_{applicant_name}"
                )
            
            comment = st.text_area(
                "Optional Comment:",
                value=current_vote['comment'] if (current_vote is not None and pd.notna(current_vote['comment'])) else "",
                key=f"comment_{applicant_name}",
                placeholder="Explain your vote..."
            )
            
            if st.button(f"💾 Submit Vote for {applicant_name}", key=f"submit_{applicant_name}"):
                original_status = current_vote['status'] if current_vote is not None else None
                original_rating = current_vote['rating'] if current_vote is not None else None
                save_vote(judge_name, applicant_name, status, rating, comment, original_status, original_rating)

# ===== TAB 2: RESULTS DASHBOARD =====
with tab2:
    st.header("📊 Voting Results")
    
    df_votes = load_votes()
    
    if df_votes.empty:
        st.info("No votes yet")
    else:
        # Get latest votes only (most recent from each judge)
        df_latest = df_votes.sort_values('timestamp').groupby(['judge_name', 'applicant_name']).tail(1)
        
        # Summary by applicant
        st.subheader("Vote Summary by Applicant")
        
        summary_data = []
        for applicant in applicant_names:
            applicant_votes = df_latest[df_latest['applicant_name'] == applicant]
            
            approve_count = (applicant_votes['status'] == 'Approve').sum()
            reject_count = (applicant_votes['status'] == 'Reject').sum()
            maybe_count = (applicant_votes['status'] == 'Maybe').sum()
            avg_rating = applicant_votes['rating'].mean()
            total_votes = len(applicant_votes)
            
            summary_data.append({
                'Applicant': applicant,
                'Approve ✅': approve_count,
                'Reject ❌': reject_count,
                'Maybe ❓': maybe_count,
                'Avg Rating ⭐': f"{avg_rating:.2f}/5" if total_votes > 0 else "N/A",
                'Total Votes': total_votes
            })
        
        df_summary = pd.DataFrame(summary_data)
        st.dataframe(df_summary, use_container_width=True, hide_index=True)
        
        st.divider()
        
        # Detailed votes by applicant
        st.subheader("Detailed Votes & Comments")
        
        selected_applicant = st.selectbox("Select Applicant:", applicant_names)
        applicant_votes = df_latest[df_latest['applicant_name'] == selected_applicant].sort_values('judge_name')
        
        if not applicant_votes.empty:
            for idx, vote in applicant_votes.iterrows():
                status_emoji = {"Approve": "✅", "Reject": "❌", "Maybe": "❓"}
                emoji = status_emoji.get(vote['status'], "")
                
                st.write(f"**{vote['judge_name']}** {emoji} {vote['status']} | ⭐ {int(vote['rating'])}/5 | _{vote['timestamp']}_")
                
                if pd.notna(vote['comment']) and vote['comment']:
                    st.write(f"> {vote['comment']}")
                
                if pd.notna(vote['original_status']) and vote['original_status']:
                    st.caption(f"🔄 Changed from: {vote['original_status']} | ⭐ {int(vote['original_rating'])}/5")
                
                st.divider()
        else:
            st.info(f"No votes for {selected_applicant} yet")
        
        # Judge summary
        st.subheader("📋 Votes by Judge")
        
        judge_summary = []
        for judge in df_latest['judge_name'].unique():
            judge_votes = df_latest[df_latest['judge_name'] == judge]
            judge_summary.append({
                'Judge': judge,
                'Votes Cast': len(judge_votes),
                'Approve': (judge_votes['status'] == 'Approve').sum(),
                'Reject': (judge_votes['status'] == 'Reject').sum(),
                'Maybe': (judge_votes['status'] == 'Maybe').sum()
            })
        
        df_judge_summary = pd.DataFrame(judge_summary)
        st.dataframe(df_judge_summary, use_container_width=True, hide_index=True)

# ===== TAB 3: EXPORT RESULTS =====
with tab3:
    st.header("📥 Export Voting Results")
    
    df_votes = load_votes()
    
    if df_votes.empty:
        st.warning("No votes to export yet")
    else:
        # Get latest votes only
        df_latest = df_votes.sort_values('timestamp').groupby(['judge_name', 'applicant_name']).tail(1)
        
        if st.button("📊 Generate Excel Summary Report"):
            # Create Excel workbook
            with pd.ExcelWriter(export_file, engine='openpyxl') as writer:
                # Sheet 1: Summary by Applicant
                summary_data = []
                for applicant in applicant_names:
                    applicant_votes = df_latest[df_latest['applicant_name'] == applicant]
                    
                    approve_count = (applicant_votes['status'] == 'Approve').sum()
                    reject_count = (applicant_votes['status'] == 'Reject').sum()
                    maybe_count = (applicant_votes['status'] == 'Maybe').sum()
                    avg_rating = applicant_votes['rating'].mean()
                    total_votes = len(applicant_votes)
                    
                    summary_data.append({
                        'Applicant': applicant,
                        'Total Votes': total_votes,
                        'Approve': approve_count,
                        'Reject': reject_count,
                        'Maybe': maybe_count,
                        'Avg Rating': f"{avg_rating:.2f}" if total_votes > 0 else "N/A"
                    })
                
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # Sheet 2: All Votes
                df_votes_export = df_latest[['timestamp', 'judge_name', 'applicant_name', 'status', 'rating', 'comment']].copy()
                df_votes_export.to_excel(writer, sheet_name='All Votes', index=False)
                
                # Sheet 3: Judge Summary
                judge_summary = []
                for judge in df_latest['judge_name'].unique():
                    judge_votes = df_latest[df_latest['judge_name'] == judge]
                    judge_summary.append({
                        'Judge': judge,
                        'Votes Cast': len(judge_votes),
                        'Approve': (judge_votes['status'] == 'Approve').sum(),
                        'Reject': (judge_votes['status'] == 'Reject').sum(),
                        'Maybe': (judge_votes['status'] == 'Maybe').sum()
                    })
                
                df_judge_summary = pd.DataFrame(judge_summary)
                df_judge_summary.to_excel(writer, sheet_name='Judge Summary', index=False)
            
            st.success(f"✅ Export complete! Saved to:\n`{export_file}`")
            st.balloons()
        
        st.divider()
        st.subheader("📋 Current Votes Preview:")
        st.dataframe(df_latest[['judge_name', 'applicant_name', 'status', 'rating', 'comment']].sort_values(['applicant_name', 'judge_name']), use_container_width=True, hide_index=True)

st.divider()
st.caption(f"Voting data stored in: {votes_file}")
